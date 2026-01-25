"""
Base Excel exporter class with styles and helper methods.
"""

from io import BytesIO
from pathlib import Path
from typing import List, Any, Dict, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from PIL import Image as PILImage
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False


class XlsxExporterBase:
    """Base class for Excel export with shared styles and helpers."""
    
    # Styles
    HEADER_FONT = Font(bold=True, color="FFFFFF") if OPENPYXL_AVAILABLE else None
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid") if OPENPYXL_AVAILABLE else None
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True) if OPENPYXL_AVAILABLE else None
    ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") if OPENPYXL_AVAILABLE else None
    THIN_BORDER = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    ) if OPENPYXL_AVAILABLE else None
    
    # Image settings
    DEFAULT_IMAGE_SIZE = 32  # pixels
    IMAGE_ROW_HEIGHT = 28  # points (roughly 32 pixels at 96 DPI)
    
    def __init__(self, output_path: str | Path):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required. Install with: pip install openpyxl")
        self.output_path = Path(output_path)
        self.output_path.mkdir(parents=True, exist_ok=True)
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        
        # Image storage - populated by set_image_extractor()
        self._object_images: Dict[int, Any] = {}  # object_id -> PIL Image
        self._npc_images: Dict[int, Any] = {}  # npc_id -> PIL Image
        self._images_available = False
    
    def set_image_extractor(self, image_extractor) -> None:
        """
        Set the image extractor to use for embedding images.
        
        Args:
            image_extractor: An ImageExtractor instance with extracted images
        """
        if image_extractor is None:
            return
        
        # Store object images (items, weapons, armor, etc.)
        if hasattr(image_extractor, 'extracted_images'):
            self._object_images = image_extractor.extracted_images.copy()
        
        # Store NPC images
        if hasattr(image_extractor, 'extracted_npc_images'):
            self._npc_images = image_extractor.extracted_npc_images.copy()
        
        self._images_available = bool(self._object_images or self._npc_images)
        if self._images_available:
            print(f"  Images available: {len(self._object_images)} objects, {len(self._npc_images)} NPCs")
    
    def _get_object_image(self, object_id: int) -> Optional[Any]:
        """Get PIL image for an object ID."""
        return self._object_images.get(object_id)
    
    def _get_npc_image(self, npc_id: int) -> Optional[Any]:
        """Get PIL image for an NPC ID."""
        return self._npc_images.get(npc_id)
    
    def _add_image_to_cell(self, ws, pil_image, cell: str, 
                          width: int = None, height: int = None) -> bool:
        """
        Add a PIL image to a worksheet cell.
        
        Args:
            ws: Worksheet to add image to
            pil_image: PIL Image object
            cell: Cell reference (e.g., 'A2')
            width: Target width in pixels (default: DEFAULT_IMAGE_SIZE)
            height: Target height in pixels (default: DEFAULT_IMAGE_SIZE)
            
        Returns:
            True if image was added successfully, False otherwise
        """
        if not PIL_AVAILABLE or pil_image is None:
            return False
        
        width = width or self.DEFAULT_IMAGE_SIZE
        height = height or self.DEFAULT_IMAGE_SIZE
        
        try:
            # Scale image while maintaining aspect ratio
            img_width, img_height = pil_image.size
            scale = min(width / img_width, height / img_height)
            new_width = int(img_width * scale)
            new_height = int(img_height * scale)
            
            # Resize if needed (use LANCZOS for high quality)
            if scale != 1.0:
                resized = pil_image.resize((new_width, new_height), PILImage.Resampling.LANCZOS)
            else:
                resized = pil_image
            
            # Convert to bytes
            img_bytes = BytesIO()
            resized.save(img_bytes, format='PNG')
            img_bytes.seek(0)
            
            # Create openpyxl image
            xl_img = XlImage(img_bytes)
            xl_img.width = new_width
            xl_img.height = new_height
            
            # Add to worksheet
            ws.add_image(xl_img, cell)
            return True
            
        except Exception as e:
            # Silently fail - missing image shouldn't break export
            return False
    
    def _set_row_height_for_image(self, ws, row_num: int) -> None:
        """Set row height to accommodate images."""
        ws.row_dimensions[row_num].height = self.IMAGE_ROW_HEIGHT
    
    def _create_sheet(self, name: str, headers: List[str]) -> Any:
        """Create a new sheet with headers."""
        ws = self.wb.create_sheet(name)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER
        ws.freeze_panes = 'A2'
        return ws
    
    def _auto_column_width(self, ws, min_width: int = 8, max_width: int = 80, 
                           skip_columns: List[str] = None):
        """Auto-adjust column widths based on content.
        
        Args:
            ws: Worksheet
            min_width: Minimum column width
            max_width: Maximum column width
            skip_columns: List of column letters to skip (e.g., ['A'] for image columns)
        """
        skip_columns = skip_columns or []
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            
            # Skip specified columns (typically image columns)
            if column in skip_columns:
                continue
            
            for cell in column_cells:
                try:
                    if cell.value:
                        cell_len = len(str(cell.value))
                        if '\n' in str(cell.value):
                            cell_len = max(len(line) for line in str(cell.value).split('\n'))
                        max_length = max(max_length, cell_len)
                except:
                    pass
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column].width = adjusted_width
    
    def _add_row(self, ws, row_num: int, values: List[Any], alternate: bool = False):
        """Add a row of values to the sheet."""
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = self.THIN_BORDER
            if alternate:
                cell.fill = self.ALT_ROW_FILL
    
    def save(self, filename: str = "ultima_underworld_data.xlsx") -> Path:
        """Save the workbook."""
        filepath = self.output_path / filename
        self.wb.save(filepath)
        return filepath
