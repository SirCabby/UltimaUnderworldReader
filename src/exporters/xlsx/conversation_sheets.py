"""
Conversation-related sheet exports for Excel exporter.
"""

from typing import Dict

try:
    from openpyxl.styles import Alignment
except ImportError:
    Alignment = None

from ...parsers.conversation_parser import Opcode


class ConversationSheetsMixin:
    """Mixin providing conversation-related sheet exports."""
    
    def export_conversations_structured(self, conversations: Dict, strings_parser, npc_names: Dict) -> None:
        """Export conversation structure with NPC dialogue and player response options."""
        headers = ["NPC Name", "Conv Slot", "String #", "Type", "Text"]
        ws = self._create_sheet("Conversations", headers)
        
        block7 = strings_parser.get_block(7) or []
        
        row = 2
        for slot in sorted(conversations.keys()):
            conv = conversations[slot]
            
            name_idx = slot + 16
            npc_name = block7[name_idx] if name_idx < len(block7) else f"NPC #{slot}"
            
            dialogue_strings = strings_parser.get_block(conv.string_block) or []
            
            # Analyze bytecode for string types
            npc_says, player_responses = self._analyze_conversation_bytecode(conv, dialogue_strings)
            
            # Output strings with their types
            for idx, text in enumerate(dialogue_strings):
                text = text.strip()
                if not text or text.startswith('@'):
                    continue
                
                if idx in npc_says and idx not in player_responses:
                    line_type = "NPC"
                elif idx in player_responses and idx not in npc_says:
                    line_type = "Player"
                elif idx in npc_says and idx in player_responses:
                    line_type = "Both"
                else:
                    line_type = "Unknown"
                
                values = [npc_name, slot, idx, line_type, text]
                self._add_row(ws, row, values, row % 2 == 0)
                row += 1
            
            # Add separator row between NPCs
            if slot != max(conversations.keys()):
                self._add_row(ws, row, ["", "", "", "", ""], False)
                row += 1
        
        self._auto_column_width(ws)
    
    def _analyze_conversation_bytecode(self, conv, dialogue_strings) -> tuple:
        """Analyze conversation bytecode to identify NPC vs player dialogue."""
        npc_says = set()
        player_responses = set()
        
        for i, instr in enumerate(conv.code):
            # SAY_OP preceded by PUSHI = NPC dialogue
            if instr.opcode == Opcode.SAY_OP:
                for j in range(i-1, max(0, i-5), -1):
                    prev = conv.code[j]
                    if prev.opcode == Opcode.PUSHI and prev.operand is not None:
                        npc_says.add(prev.operand)
                        break
            
            # CALLI 0 (babl_menu) - strings pushed before it are player responses
            elif instr.opcode == Opcode.CALLI and instr.operand == 0:
                for j in range(i-1, max(0, i-40), -1):
                    prev = conv.code[j]
                    if prev.opcode == Opcode.PUSHI and prev.operand is not None:
                        if prev.operand < len(dialogue_strings):
                            player_responses.add(prev.operand)
                    if prev.opcode in (Opcode.SAY_OP, Opcode.CALLI, Opcode.ADDSP):
                        break
        
        return npc_says, player_responses
    
    def export_conversations_full(self, conversations: Dict, strings_parser, npc_names: Dict) -> None:
        """Export complete conversation blocks for each NPC."""
        headers = ["NPC Name", "Conv Slot", "String Block", "Full Dialogue"]
        ws = self._create_sheet("Full Dialogues", headers)
        
        block7 = strings_parser.get_block(7) or []
        
        row = 2
        for slot in sorted(conversations.keys()):
            conv = conversations[slot]
            
            name_idx = slot + 16
            npc_name = block7[name_idx] if name_idx < len(block7) else f"NPC #{slot}"
            
            dialogue_strings = strings_parser.get_block(conv.string_block) or []
            
            # Build complete dialogue
            lines = []
            for s in dialogue_strings:
                s = s.strip()
                if s and not s.startswith('@'):
                    lines.append(s)
            
            full_dialogue = "\n".join(lines)
            
            values = [npc_name, slot, f"0x{conv.string_block:04X}", full_dialogue]
            self._add_row(ws, row, values, row % 2 == 0)
            
            # Set row height for long text
            ws.row_dimensions[row].height = min(400, max(15, len(lines) * 12))
            
            row += 1
        
        # Wrap text in dialogue column
        if Alignment:
            for cell in ws['D']:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        self._auto_column_width(ws)
    
    def export_dialogue_responses(self, conversations: Dict, strings_parser, npc_names: Dict) -> None:
        """Legacy method - redirects to structured export."""
        pass
