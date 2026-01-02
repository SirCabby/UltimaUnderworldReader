"""
Excel (XLSX) Exporter for Ultima Underworld extracted data.

Exports all extracted game data to a multi-sheet Excel workbook.
"""

from pathlib import Path
from typing import Dict, List, Any, Optional, Set
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from ..constants import (
    COMPLETE_MANTRAS,
    NPC_GOALS,
    NPC_ONLY_SPELLS,
    PLAYER_SPELLS,
    STACKABLE_ITEMS,
    CARRYABLE_CATEGORIES,
    CATEGORY_DISPLAY_NAMES,
    CARRYABLE_CONTAINERS,
    NPC_ATTITUDES,
    RUNE_MEANINGS,
)
from ..utils import parse_item_name
from ..parsers.conversation_parser import Opcode


class XlsxExporter:
    """Exports game data to Excel xlsx format."""
    
    # Styles
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    def __init__(self, output_path: str | Path):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required. Install with: pip install openpyxl")
        self.output_path = Path(output_path)
        self.output_path.mkdir(parents=True, exist_ok=True)
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
    
    def _create_sheet(self, name: str, headers: List[str]) -> Any:
        ws = self.wb.create_sheet(name)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER
        ws.freeze_panes = 'A2'
        return ws
    
    def _auto_column_width(self, ws, min_width: int = 8, max_width: int = 80):
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value:
                        cell_len = len(str(cell.value))
                        if '\n' in str(cell.value):
                            cell_len = max(len(line) for line in str(cell.value).split('\n'))
                        max_length = max(max_length, cell_len)
                except:
                    pass
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column].width = adjusted_width
    
    def _add_row(self, ws, row_num: int, values: List[Any], alternate: bool = False):
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = self.THIN_BORDER
            if alternate:
                cell.fill = self.ALT_ROW_FILL
    
    def export_items(self, item_types: Dict, placed_items: List) -> None:
        """Export all item types."""
        headers = [
            "ID", "ID (Hex)", "Name", "Category",
            "Weight", "Value", 
            "Property 1", "Property 2", "Property 3", "Property 4"
        ]
        ws = self._create_sheet("Items", headers)
        
        row = 2
        for item_id in sorted(item_types.keys()):
            item = item_types[item_id]
            props = item.properties
            prop_strs = [f"{k}: {v}" for k, v in list(props.items())[:4]]
            prop_strs.extend([""] * (4 - len(prop_strs)))
            
            # Determine if item can be carried based on category
            is_carryable = item.category in CARRYABLE_CATEGORIES
            
            # Weight in stones (mass is stored in 0.1 stone units)
            # Only show weight if it's actually defined (> 0)
            if is_carryable and item.mass > 0:
                weight_str = f"{item.mass / 10:.1f}"
            else:
                weight_str = ""
            
            # Value in gold (value is already in whole gold pieces)
            value_str = str(item.value) if item.value > 0 else ""
            
            values = [
                item.item_id, f"0x{item.item_id:03X}", item.name, item.category,
                weight_str, value_str,
                prop_strs[0], prop_strs[1], prop_strs[2], prop_strs[3]
            ]
            self._add_row(ws, row, values, row % 2 == 0)
            row += 1
        self._auto_column_width(ws)
    
    def export_weapons(self, item_types: Dict, objects_parser) -> None:
        """Export weapons."""
        headers = ["ID", "Name", "Type", "Slash", "Bash", "Stab", "Skill", "Durability", "Weight", "Value"]
        ws = self._create_sheet("Weapons", headers)
        
        row = 2
        for item_id in range(0x10):
            item = item_types.get(item_id)
            weapon = objects_parser.get_melee_weapon(item_id)
            if item and weapon:
                skill = weapon.skill_type.name if hasattr(weapon.skill_type, 'name') else str(weapon.skill_type)
                # Weight in stones (mass is in 0.1 stones)
                weight_str = f"{item.mass / 10:.1f}" if item.mass > 0 else ""
                values = [item_id, item.name, "Melee", weapon.slash_damage, weapon.bash_damage,
                         weapon.stab_damage, skill, weapon.durability, weight_str, item.value]
                self._add_row(ws, row, values, row % 2 == 0)
                row += 1
        
        for item_id in range(0x10, 0x20):
            item = item_types.get(item_id)
            weapon = objects_parser.get_ranged_weapon(item_id)
            if item and weapon:
                # Weight in stones (mass is in 0.1 stones)
                weight_str = f"{item.mass / 10:.1f}" if item.mass > 0 else ""
                values = [item_id, item.name, "Ranged", "-", "-", "-", "Missile",
                         weapon.durability, weight_str, item.value]
                self._add_row(ws, row, values, row % 2 == 0)
                row += 1
        self._auto_column_width(ws)
    
    def export_armor(self, item_types: Dict, objects_parser) -> None:
        """Export armor."""
        headers = ["ID", "Name", "Category", "Protection", "Durability", "Weight", "Value"]
        ws = self._create_sheet("Armor", headers)
        
        row = 2
        for item_id in range(0x20, 0x40):
            item = item_types.get(item_id)
            armor = objects_parser.get_armour(item_id)
            if item and armor:
                cat = armor.category.name if hasattr(armor.category, 'name') else str(armor.category)
                # Weight in stones (mass is in 0.1 stones)
                weight_str = f"{item.mass / 10:.1f}" if item.mass > 0 else ""
                values = [item_id, item.name, cat, armor.protection, armor.durability, weight_str, item.value]
                self._add_row(ws, row, values, row % 2 == 0)
                row += 1
        self._auto_column_width(ws)
    
    def export_npcs(self, npcs: List, npc_names: Dict, strings_parser, level_parser=None) -> None:
        """Export NPCs with correct name mapping, goal descriptions, and inventory."""
        headers = [
            "Level", "Tile X", "Tile Y", "Creature Type", "Creature ID",
            "Named NPC", "Conv Slot", "HP",
            "Attitude", "Goal", "Goal Description", "Inventory", "Home X", "Home Y"
        ]
        ws = self._create_sheet("NPCs", headers)
        
        obj_names = strings_parser.get_block(4) or []
        block7 = strings_parser.get_block(7) or []
        
        # Filter to only NPCs that are actually placed (have valid tile coords)
        placed_npcs = [n for n in npcs if n.tile_x > 0 or n.tile_y > 0]
        sorted_npcs = sorted(placed_npcs, key=lambda n: (n.level, n.tile_x, n.tile_y))
        
        row = 2
        for npc in sorted_npcs:
            # Get creature type from object names
            creature_type = ""
            if npc.object_id < len(obj_names):
                raw = obj_names[npc.object_id]
                if raw:
                    creature_type, _, _ = parse_item_name(raw)
            
            # Get named NPC from conversation slot (whoami + 16 indexes into block 7)
            # IMPORTANT: Only use this for NPCs with conversation_slot > 0
            named_npc = ""
            if npc.conversation_slot > 0:
                name_idx = npc.conversation_slot + 16
                if name_idx < len(block7) and block7[name_idx]:
                    named_npc = block7[name_idx]
            
            # Get goal description
            goal_desc = NPC_GOALS.get(npc.goal, f"Unknown ({npc.goal})")
            attitude_name = NPC_ATTITUDES.get(npc.attitude, str(npc.attitude))
            
            # Get NPC inventory if level_parser available
            inventory_str = ""
            if level_parser and hasattr(npc, 'special_link') and npc.special_link > 0:
                level = level_parser.get_level(npc.level)
                if level:
                    spell_names = strings_parser.get_block(6) or []
                    block5 = strings_parser.get_block(5) or []
                    inv_items = []
                    current = npc.special_link
                    seen = set()
                    while current > 0 and current not in seen and len(inv_items) < 20:
                        seen.add(current)
                        if current in level.objects:
                            obj = level.objects[current]
                            item_name, _, _ = parse_item_name(obj_names[obj.item_id] if obj.item_id < len(obj_names) else "")
                            
                            # Add metadata for special items
                            if item_name:
                                item_desc = item_name
                                
                                # Keys - show description from block5[100 + owner]
                                if 0x100 <= obj.item_id <= 0x10E:
                                    if obj.owner > 0:
                                        desc_idx = 100 + obj.owner
                                        if desc_idx < len(block5) and block5[desc_idx]:
                                            key_desc = block5[desc_idx]
                                            item_desc = f"key ({key_desc[:40]}...)" if len(key_desc) > 40 else f"key ({key_desc})"
                                        else:
                                            item_desc = f"key (lock #{obj.owner})"
                                
                                # Scrolls - readable text
                                elif 0x138 <= obj.item_id <= 0x13F and obj.item_id != 0x13B:
                                    link = obj.quantity_or_link
                                    if obj.is_quantity and link >= 512:
                                        item_desc = f"scroll (text #{link-512})"
                                    else:
                                        item_desc = "scroll"
                                
                                # Books
                                elif 0x130 <= obj.item_id <= 0x137:
                                    link = obj.quantity_or_link
                                    if obj.is_quantity and link >= 512:
                                        item_desc = f"book (text #{link-512})"
                                    else:
                                        item_desc = "book"
                                
                                # Potions
                                elif obj.item_id == 0xBB:
                                    item_desc = "red potion (mana)"
                                elif obj.item_id == 0xBC:
                                    item_desc = "green potion (heal)"
                                
                                # Wands - show spell from linked object
                                elif 0x98 <= obj.item_id <= 0x9B:
                                    if not obj.is_quantity:
                                        link = obj.quantity_or_link
                                        if link in level.objects:
                                            spell_obj = level.objects[link]
                                            if spell_obj.item_id == 0x120:
                                                spell_idx = spell_obj.quality + 256 if spell_obj.quality < 64 else spell_obj.quality
                                                spell = spell_names[spell_idx] if spell_idx < len(spell_names) else ""
                                                if spell:
                                                    item_desc = f"wand of {spell} ({obj.quality} charges)"
                                                else:
                                                    item_desc = f"wand ({obj.quality} charges)"
                                            else:
                                                item_desc = f"wand ({obj.quality} charges)"
                                        else:
                                            item_desc = f"wand ({obj.quality} charges)"
                                    else:
                                        item_desc = f"wand ({obj.quality} charges)"
                                
                                inv_items.append(item_desc)
                            current = obj.next_index
                        else:
                            break
                    inventory_str = ", ".join(inv_items) if inv_items else ""
            
            values = [
                npc.level + 1, npc.tile_x, npc.tile_y, creature_type, f"0x{npc.object_id:02X}",
                named_npc, npc.conversation_slot if npc.conversation_slot > 0 else "",
                npc.hp,
                attitude_name, npc.goal, goal_desc, inventory_str,
                npc.home_x, npc.home_y
            ]
            self._add_row(ws, row, values, row % 2 == 0)
            row += 1
        self._auto_column_width(ws)
    
    def export_spells(self, spells: List, spell_runes: Dict) -> None:
        """Export spells with descriptions and caster info."""
        from ..constants import SPELL_DESCRIPTIONS
        
        headers = ["ID", "Name", "Circle", "Rune Combination", "Description", "Caster"]
        ws = self._create_sheet("Spells", headers)
        
        row = 2
        for spell in spells:
            runes = spell_runes.get(spell.name, [])
            rune_str = " ".join(runes) if runes else ""
            
            # Get description
            description = SPELL_DESCRIPTIONS.get(spell.name, "")
            
            # Only show circle for player spells that have rune combinations
            circle_str = spell.circle if rune_str else ""
            
            if spell.name in NPC_ONLY_SPELLS:
                caster = "NPC Only"
            elif spell.name in PLAYER_SPELLS or rune_str:
                caster = "Player"
            else:
                caster = ""
            
            values = [spell.spell_id, spell.name, circle_str, rune_str, description, caster]
            self._add_row(ws, row, values, row % 2 == 0)
            row += 1
        self._auto_column_width(ws)
    
    def export_runes(self, runes: Dict) -> None:
        """Export runes."""
        headers = ["ID", "Rune Name", "Meaning"]
        ws = self._create_sheet("Runes", headers)
        
        row = 2
        for rune_id in sorted(runes.keys()):
            name = runes[rune_id]
            meaning = RUNE_MEANINGS.get(name, "")
            self._add_row(ws, row, [rune_id, name, meaning], row % 2 == 0)
            row += 1
        self._auto_column_width(ws)
    
    def export_mantras(self) -> None:
        """Export complete mantra list from game data with point increases."""
        headers = ["Mantra", "Skill(s) Affected", "Effect/Notes", "Point Increase"]
        ws = self._create_sheet("Mantras", headers)
        
        row = 2
        for mantra, skills, notes, points in COMPLETE_MANTRAS:
            self._add_row(ws, row, [mantra, skills, notes, points], row % 2 == 0)
            row += 1
        self._auto_column_width(ws)
    
    def export_conversations_structured(self, conversations: Dict, strings_parser, npc_names: Dict) -> None:
        """Export conversation structure with NPC dialogue and player response options."""
        headers = ["NPC Name", "Conv Slot", "String #", "Type", "Text"]
        ws = self._create_sheet("Conversations", headers)
        
        block7 = strings_parser.get_block(7) or []
        
        row = 2
        for slot in sorted(conversations.keys()):
            conv = conversations[slot]
            
            name_idx = slot + 16
            npc_name = block7[name_idx] if name_idx < len(block7) else f"NPC #{slot}"
            
            dialogue_strings = strings_parser.get_block(conv.string_block) or []
            
            # Analyze bytecode to find which strings are NPC dialogue (SAY_OP)
            # and which are player responses (pushed before babl_menu/CALLI 0)
            npc_says: set = set()
            player_responses: set = set()
            
            for i, instr in enumerate(conv.code):
                # SAY_OP preceded by PUSHI = NPC dialogue
                if instr.opcode == Opcode.SAY_OP:
                    # Look back for the PUSHI with the string index
                    for j in range(i-1, max(0, i-5), -1):
                        prev = conv.code[j]
                        if prev.opcode == Opcode.PUSHI and prev.operand is not None:
                            npc_says.add(prev.operand)
                            break
                
                # CALLI 0 (babl_menu) - strings pushed before it are player responses
                elif instr.opcode == Opcode.CALLI and instr.operand == 0:
                    # Look back for PUSHI instructions
                    for j in range(i-1, max(0, i-40), -1):
                        prev = conv.code[j]
                        if prev.opcode == Opcode.PUSHI and prev.operand is not None:
                            if prev.operand < len(dialogue_strings):
                                player_responses.add(prev.operand)
                        # Stop at certain opcodes that indicate end of menu setup
                        if prev.opcode in (Opcode.SAY_OP, Opcode.CALLI, Opcode.ADDSP):
                            break
            
            # Output strings with their types
            for idx, text in enumerate(dialogue_strings):
                text = text.strip()
                if not text or text.startswith('@'):
                    continue
                
                # Determine type based on bytecode analysis
                if idx in npc_says and idx not in player_responses:
                    line_type = "NPC"
                elif idx in player_responses and idx not in npc_says:
                    line_type = "Player"
                elif idx in npc_says and idx in player_responses:
                    line_type = "Both"  # Used in both contexts
                else:
                    line_type = "Unknown"
                
                values = [npc_name, slot, idx, line_type, text]
                self._add_row(ws, row, values, row % 2 == 0)
                row += 1
            
            # Add separator row between NPCs
            if slot != max(conversations.keys()):
                self._add_row(ws, row, ["", "", "", "", ""], False)
                row += 1
        
        self._auto_column_width(ws)
    
    def export_conversations_full(self, conversations: Dict, strings_parser, npc_names: Dict) -> None:
        """Export COMPLETE conversation blocks for each NPC (all text combined)."""
        headers = ["NPC Name", "Conv Slot", "String Block", "Full Dialogue"]
        ws = self._create_sheet("Full Dialogues", headers)
        
        block7 = strings_parser.get_block(7) or []
        
        row = 2
        for slot in sorted(conversations.keys()):
            conv = conversations[slot]
            
            # Get NPC name
            name_idx = slot + 16
            npc_name = block7[name_idx] if name_idx < len(block7) else f"NPC #{slot}"
            
            # Get ALL dialogue strings
            dialogue_strings = strings_parser.get_block(conv.string_block) or []
            
            # Build complete dialogue, filtering out code markers
            lines = []
            for s in dialogue_strings:
                s = s.strip()
                if s and not s.startswith('@'):
                    lines.append(s)
            
            full_dialogue = "\n".join(lines)
            
            values = [npc_name, slot, f"0x{conv.string_block:04X}", full_dialogue]
            self._add_row(ws, row, values, row % 2 == 0)
            
            # Set row height for long text
            ws.row_dimensions[row].height = min(400, max(15, len(lines) * 12))
            
            row += 1
        
        # Wrap text in dialogue column
        for cell in ws['D']:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        self._auto_column_width(ws)
    
    def export_dialogue_responses(self, conversations: Dict, strings_parser, npc_names: Dict) -> None:
        """Legacy method - now redirects to structured export."""
        # This is now handled by export_conversations_structured
        pass
    
    def export_placed_objects(self, placed_items: List, item_types: Dict, 
                             strings_parser, level_parser=None) -> None:
        """Export placed objects with actual locations, descriptions, and effects."""
        headers = [
            "Level", "Tile X", "Tile Y", "Item Name", "Item ID",
            "Category", "Description", "Enchantment/Effect"
        ]
        ws = self._create_sheet("Placed Objects", headers)
        
        obj_names = strings_parser.get_block(4) or []
        block3 = strings_parser.get_block(3) or []  # Book/scroll text
        block5 = strings_parser.get_block(5) or []  # Quality state descriptions
        spell_names = strings_parser.get_block(6) or []
        
        # Filter to objects that have actual tile positions
        placed_with_coords = [i for i in placed_items if i.tile_x > 0 or i.tile_y > 0]
        sorted_items = sorted(placed_with_coords, key=lambda x: (x.level, x.tile_x, x.tile_y))
        
        row = 2
        for item in sorted_items:
            # Skip NPCs, doors, triggers, traps
            if 0x40 <= item.object_id <= 0x7F:  # NPCs
                continue
            if 0x140 <= item.object_id <= 0x17F:  # Doors
                continue
            if 0x180 <= item.object_id <= 0x1BF:  # Traps/triggers
                continue
            
            # Get name
            name, _, _ = parse_item_name(obj_names[item.object_id] if item.object_id < len(obj_names) else "")
            
            # Get proper category name - prefer detailed_category if available
            cat_raw = getattr(item, 'detailed_category', None) or item.object_class
            cat_raw = cat_raw if isinstance(cat_raw, str) else ""
            category = CATEGORY_DISPLAY_NAMES.get(cat_raw, cat_raw.replace('_', ' ').title() if cat_raw else "Unknown")
            
            # Get description based on item type
            description = self._get_item_description(item, block3, block5, spell_names, level_parser)
            
            # Get effect/enchantment with spell descriptions
            effect = self._get_item_effect(item, strings_parser, level_parser)
            
            values = [
                item.level + 1, item.tile_x, item.tile_y, name, f"0x{item.object_id:03X}",
                category, description, effect
            ]
            self._add_row(ws, row, values, row % 2 == 0)
            row += 1
        self._auto_column_width(ws)
    
    def _get_item_description(self, item, block3, block5, spell_names, level_parser=None) -> str:
        """Get item description based on type and quality/owner fields."""
        object_id = item.object_id
        
        # Get the link value
        link_value = item.quantity if item.is_quantity else item.special_link
        
        # Keys (0x100-0x10E)
        if 0x100 <= object_id <= 0x10E:
            if item.owner > 0:
                desc_idx = 100 + item.owner
                if desc_idx < len(block5) and block5[desc_idx]:
                    return block5[desc_idx]
            if object_id == 0x101:
                return "A lockpick"
            return ""
        
        # Books (0x130-0x137)
        if 0x130 <= object_id <= 0x137:
            if item.is_quantity and link_value >= 512:
                text_idx = link_value - 512
                if text_idx < len(block3) and block3[text_idx]:
                    return block3[text_idx].strip()
            return ""
        
        # Scrolls (0x138-0x13F, except 0x13B map)
        if 0x138 <= object_id <= 0x13F and object_id != 0x13B:
            if item.is_quantity and link_value >= 512:
                text_idx = link_value - 512
                if text_idx < len(block3) and block3[text_idx]:
                    return block3[text_idx].strip()
            return ""
        
        # Wands (0x98-0x9B)
        if 0x98 <= object_id <= 0x9B:
            if level_parser and not item.is_quantity:
                level = level_parser.get_level(item.level)
                if level and item.special_link in level.objects:
                    spell_obj = level.objects[item.special_link]
                    if spell_obj.item_id == 0x120:
                        spell_idx = spell_obj.quality + 256 if spell_obj.quality < 64 else spell_obj.quality
                        if spell_idx < len(spell_names) and spell_names[spell_idx]:
                            return f"Wand of {spell_names[spell_idx]}"
            # Check for special wands with unique spells
            from ..constants import get_special_wand_info
            special_wand = get_special_wand_info(item.level, item.tile_x, item.tile_y)
            if special_wand:
                return f"{special_wand['name']} ({item.quality} charges)"
            return f"Wand (unknown spell, {item.quality} charges)"
        
        # Map (0x13B)
        if object_id == 0x13B:
            return "Shows explored areas"
        
        # Potions (0xBB = red, 0xBC = green)
        if object_id in (0xBB, 0xBC):
            if item.is_quantity:
                link = item.quantity
                if link >= 512:
                    ench_idx = link - 512
                    spell = ""
                    if ench_idx + 256 < len(spell_names) and spell_names[ench_idx + 256]:
                        spell = spell_names[ench_idx + 256]
                    if not spell and ench_idx < len(spell_names) and spell_names[ench_idx]:
                        spell = spell_names[ench_idx]
                    if spell:
                        return f"Potion of {spell}"
                    else:
                        return f"Potion (unknown effect #{ench_idx})"
            return ""
        
        # Coins
        if object_id == 0xA0:
            if item.is_quantity:
                return f"{item.quantity} gold pieces"
            return "Gold coin"
        
        # Arrows/bolts
        if object_id in (0x10, 0x11, 0x12):
            if item.is_quantity:
                return f"Stack of {item.quantity}"
            return ""
        
        # Runes
        if 0xE0 <= object_id <= 0xFF:
            return ""
        
        # Rings
        if object_id in (0x36, 0x38, 0x39, 0x3A):
            if item.is_enchanted and item.is_quantity:
                link = item.quantity
                if link >= 512:
                    ench_idx = link - 512
                    spell = spell_names[ench_idx] if ench_idx < len(spell_names) else ""
                    if spell:
                        return f"Ring of {spell}"
            return ""
        
        # Scenery
        if 0xC0 <= object_id <= 0xDF:
            return ""
        if object_id in (0x125,):
            return ""
        
        return self._get_quality_description(object_id, item.quality, block5)
    
    def _get_quality_description(self, object_id: int, quality: int, block5: list) -> str:
        """Get the proper quality description based on item type."""
        offset = min(5, quality // 10)
        skip_descriptions = {'massive', 'sturdy', 'new', 'smooth'}
        
        base = None
        
        # Melee weapons
        if 0x00 <= object_id <= 0x0F:
            base = 6
        # Ranged weapons
        elif 0x10 <= object_id <= 0x1F:
            if object_id in (0x10, 0x11, 0x12):
                return ""
            base = 6
        # Armor
        elif 0x20 <= object_id <= 0x3F:
            if 0x3C <= object_id <= 0x3F:
                base = 6
            elif object_id in (0x24, 0x25, 0x26, 0x27):
                base = 78
            else:
                base = 6
        # Light sources
        elif 0x90 <= object_id <= 0x97:
            base = 60
        # Wands
        elif 0x98 <= object_id <= 0x9B:
            return ""
        # Treasure
        elif 0xA0 <= object_id <= 0xAF:
            if object_id == 0xA0:
                return ""
            elif object_id in (0xA2, 0xA3, 0xA4, 0xA5, 0xA6, 0xA7):
                base = 42
            else:
                base = 36
        # Food and drinks
        elif 0xB0 <= object_id <= 0xBF:
            if 0xB0 <= object_id <= 0xB7:
                base = 18
            else:
                base = 24
        # Containers
        elif 0x80 <= object_id <= 0x8F:
            if quality >= 40:
                return ""
            base = 72
        # Books and scrolls
        elif 0x130 <= object_id <= 0x13F:
            return ""
        # Quest items and misc
        elif 0x110 <= object_id <= 0x12F:
            return ""
        # Keys
        elif 0x100 <= object_id <= 0x10F:
            return ""
        
        if base is None:
            return ""
        
        desc_idx = base + offset
        if desc_idx < len(block5) and block5[desc_idx]:
            desc = block5[desc_idx]
            if desc.lower() in skip_descriptions:
                return ""
            return desc
        
        return ""
    
    def _get_item_effect(self, item, strings_parser, level_parser=None) -> str:
        """Get enchantment/effect description for an item."""
        from ..constants import SPELL_DESCRIPTIONS, get_special_wand_info
        
        spell_names = {}
        if strings_parser:
            block6 = strings_parser.get_block(6) or []
            for i, name in enumerate(block6):
                if name and name.strip():
                    spell_names[i] = name.strip()
        
        object_id = item.object_id
        link_value = item.quantity if item.is_quantity else item.special_link
        
        def format_spell_with_description(spell_name: str) -> str:
            """Format spell name with description if available."""
            if not spell_name:
                return ""
            desc = SPELL_DESCRIPTIONS.get(spell_name, "")
            if desc:
                return f"{spell_name} ({desc})"
            return spell_name
        
        # Wands
        if 0x98 <= object_id <= 0x9B:
            if level_parser and not item.is_quantity:
                level = level_parser.get_level(item.level)
                if level and item.special_link in level.objects:
                    spell_obj = level.objects[item.special_link]
                    if spell_obj.item_id == 0x120:
                        spell_idx = spell_obj.quality + 256 if spell_obj.quality < 64 else spell_obj.quality
                        spell = spell_names.get(spell_idx, "")
                        if spell:
                            spell_with_desc = format_spell_with_description(spell)
                            return f"{spell_with_desc} ({item.quality} charges)"
            # Check for special wands with unique spells not in the spell table
            special_wand = get_special_wand_info(item.level, item.tile_x, item.tile_y)
            if special_wand:
                return f"{special_wand['name']} ({item.quality} charges)"
            return f"Unknown spell ({item.quality} charges)" if item.quality > 0 else "Empty"
        
        # Keys
        if 0x100 <= object_id <= 0x10E:
            if item.owner > 0:
                return f"Opens lock #{item.owner}"
            return ""
        
        # Books/Scrolls
        if 0x130 <= object_id <= 0x13F and object_id != 0x13B:
            if item.is_quantity and link_value >= 512:
                return f"Text #{link_value - 512}"
            return ""
        
        # Potions
        if object_id in (0xBB, 0xBC):
            if item.is_quantity and link_value >= 512:
                raw_idx = link_value - 512
                spell_256 = spell_names.get(raw_idx + 256, "")
                if spell_256:
                    return format_spell_with_description(spell_256)
                spell_raw = spell_names.get(raw_idx, "")
                if spell_raw:
                    return format_spell_with_description(spell_raw)
                return f"Effect #{raw_idx}"
            if object_id == 0xBB:
                return "Restores Mana"
            else:
                return "Heals Wounds"
        
        if not item.is_enchanted:
            return ""
        
        if item.is_quantity:
            link = item.quantity
        else:
            link = item.special_link
        
        if link >= 512:
            ench_property = link - 512
        else:
            return ""
        
        # Weapons
        if object_id < 0x20:
            if 192 <= ench_property <= 199:
                spell_idx = 448 + (ench_property - 192)
                spell = spell_names.get(spell_idx, "")
                if spell:
                    return format_spell_with_description(spell)
                return f"Accuracy +{ench_property - 191}"
            elif 200 <= ench_property <= 207:
                spell_idx = 456 + (ench_property - 200)
                spell = spell_names.get(spell_idx, "")
                if spell:
                    return format_spell_with_description(spell)
                return f"Damage +{ench_property - 199}"
            elif ench_property < 64:
                spell_idx = 256 + ench_property
                spell = spell_names.get(spell_idx, "")
                return format_spell_with_description(spell)
            else:
                # Values 64-191: Look up directly in spell names
                # This includes Cursed (144-159), various spell effects, etc.
                spell = spell_names.get(ench_property, "")
                if spell:
                    return format_spell_with_description(spell)
                return f"Enchantment #{ench_property}"
        
        # Rings
        elif object_id in (0x36, 0x38, 0x39, 0x3A):
            spell = spell_names.get(ench_property, "")
            if spell:
                return format_spell_with_description(spell)
            return f"Unknown enchantment ({ench_property})"
        
        # Armor
        elif 0x20 <= object_id < 0x40:
            if 192 <= ench_property <= 199:
                spell_idx = 464 + (ench_property - 192)
                spell = spell_names.get(spell_idx, "")
                if spell:
                    return format_spell_with_description(spell)
                return f"Protection +{ench_property - 191}"
            elif 200 <= ench_property <= 207:
                spell_idx = 472 + (ench_property - 200)
                spell = spell_names.get(spell_idx, "")
                if spell:
                    return format_spell_with_description(spell)
                return f"Toughness +{ench_property - 199}"
            elif ench_property < 64:
                spell_idx = 256 + ench_property
                spell = spell_names.get(spell_idx, "")
                return format_spell_with_description(spell)
            else:
                # Values 64-191: Look up directly in spell names
                # This includes Cursed (144-159), various spell effects, etc.
                spell = spell_names.get(ench_property, "")
                if spell:
                    return format_spell_with_description(spell)
                return f"Enchantment #{ench_property}"
        
        return ""
    
    def export_unused_items(self, item_types: Dict, placed_items: List, strings_parser) -> None:
        """Export items never placed in game."""
        headers = ["ID", "ID (Hex)", "Name", "Category", "Notes"]
        ws = self._create_sheet("Unused Items", headers)
        
        obj_names = strings_parser.get_block(4) or []
        placed_ids: Set[int] = {item.object_id for item in placed_items}
        
        row = 2
        for item_id in sorted(item_types.keys()):
            if item_id not in placed_ids:
                item = item_types[item_id]
                
                name, _, _ = parse_item_name(obj_names[item_id] if item_id < len(obj_names) else "")
                if not name:
                    name = "(unnamed)"
                
                notes = ""
                if 0x1C0 <= item_id <= 0x1FF:
                    notes = "System object"
                elif not name or name == "(unnamed)":
                    notes = "Unused/padding"
                else:
                    notes = "Defined but never placed"
                
                values = [item_id, f"0x{item_id:03X}", name, item.category, notes]
                self._add_row(ws, row, values, row % 2 == 0)
                row += 1
        self._auto_column_width(ws)
    
    def export_secrets(self, secrets: List) -> None:
        """Export secrets."""
        headers = ["Level", "Type", "Tile X", "Tile Y", "Object ID", "Description"]
        ws = self._create_sheet("Secrets & Traps", headers)
        
        sorted_secrets = sorted(secrets, key=lambda s: (s.level, s.secret_type, s.tile_x, s.tile_y))
        
        row = 2
        for secret in sorted_secrets:
            values = [
                secret.level + 1, secret.secret_type, secret.tile_x, secret.tile_y,
                f"0x{secret.object_id:03X}" if secret.object_id else "",
                secret.description
            ]
            self._add_row(ws, row, values, row % 2 == 0)
            row += 1
        self._auto_column_width(ws)
    
    def export_containers(self, item_types: Dict, objects_parser, common_parser=None) -> None:
        """Export carryable containers only."""
        headers = ["ID", "Name", "Weight", "Capacity (stones)", "Accepts"]
        ws = self._create_sheet("Containers", headers)
        
        row = 2
        for item_id in sorted(CARRYABLE_CONTAINERS.keys()):
            item = item_types.get(item_id)
            container = objects_parser.get_container(item_id)
            if item and container:
                # Weight in stones (mass is in 0.1 stones)
                weight_str = f"{item.mass / 10:.1f}" if item.mass > 0 else ""
                
                values = [
                    item_id, item.name, weight_str, container.capacity_stones,
                    container.accepted_type_name
                ]
                self._add_row(ws, row, values, row % 2 == 0)
                row += 1
        self._auto_column_width(ws)
    
    def export_food(self, item_types: Dict, strings_parser=None) -> None:
        """Export food items with nutrition values."""
        from ..constants import FOOD_NUTRITION, FOOD_NOTES, FOOD_ID_MIN, FOOD_ID_MAX
        
        headers = ["ID", "Name", "Nutrition", "Weight", "Nutrition/Weight", "Notes"]
        ws = self._create_sheet("Food", headers)
        
        # Get object names from strings
        obj_names = []
        if strings_parser:
            obj_names = strings_parser.get_block(4) or []
        
        row = 2
        for item_id in range(FOOD_ID_MIN, FOOD_ID_MAX + 1):
            item = item_types.get(item_id)
            nutrition = FOOD_NUTRITION.get(item_id, 0)
            notes = FOOD_NOTES.get(item_id, "")
            
            # Get name from strings or item_types
            name = ""
            if item:
                name = item.name
            elif item_id < len(obj_names) and obj_names[item_id]:
                name, _, _ = parse_item_name(obj_names[item_id])
            
            # Weight in stones (mass is in 0.1 stones)
            weight = 0.0
            weight_str = ""
            if item and item.mass > 0:
                weight = item.mass / 10.0
                weight_str = f"{weight:.1f}"
            
            # Calculate nutrition per weight (efficiency)
            efficiency_str = ""
            if weight > 0:
                efficiency = nutrition / weight
                efficiency_str = f"{efficiency:.1f}"
            
            values = [
                item_id, name, nutrition, weight_str, efficiency_str, notes
            ]
            self._add_row(ws, row, values, row % 2 == 0)
            row += 1
        self._auto_column_width(ws)
    
    def export_light_sources(self, item_types: Dict, objects_parser, strings_parser=None) -> None:
        """Export light sources (lit items only) and light-granting spells."""
        from ..constants import SPELL_CIRCLES, LIGHT_SPELL_LEVELS
        
        headers = ["Type", "Name", "Light Level", "Duration", "Notes"]
        ws = self._create_sheet("Light Sources", headers)
        
        # Get object names from strings
        obj_names = []
        if strings_parser:
            obj_names = strings_parser.get_block(4) or []
        
        row = 2
        
        # Duration conversion: game duration units to approximate real-time seconds
        # Based on UW game timing: each duration unit ≈ 5 minutes of gameplay
        DURATION_SECONDS = {
            1: 300,   # ~5 minutes (candle burns fast)
            2: 600,   # ~10 minutes (torch)
            3: 900,   # ~15 minutes (but taper is special - eternal in practice)
            4: 1200,  # ~20 minutes (lantern lasts longest)
        }
        
        # Item brightness to normalized light level (0-7 scale like spells)
        # Raw data: lantern=10, torch=3, candle=12 (these are quality/burn values, not light)
        # Actual in-game brightness ranking: lantern > torch > candle > taper (magical)
        ITEM_LIGHT_LEVELS = {
            0x94: 5,  # Lit lantern - brightest portable
            0x95: 4,  # Lit torch - standard light
            0x96: 2,  # Lit candle - dim
            0x97: 3,  # Lit taper - magical light (Taper of Sacrifice)
        }
        
        # Light source items from OBJECTS.DAT - only lit versions
        for obj_id in (0x94, 0x95, 0x96, 0x97):  # Lit lantern, torch, candle, taper
            light = objects_parser.get_light_source(obj_id)
            if light:
                # Get item name from strings
                name = ""
                if obj_id < len(obj_names) and obj_names[obj_id]:
                    name, _, _ = parse_item_name(obj_names[obj_id])
                if not name:
                    name = f"Light Source 0x{obj_id:02X}"
                
                # Get normalized light level
                light_level = ITEM_LIGHT_LEVELS.get(obj_id, 3)
                
                # Format duration in seconds
                if obj_id == 0x97:  # Lit taper (Taper of Sacrifice is eternal)
                    duration_str = "Eternal"
                    notes = "Magical item from Shrine of Spirituality"
                else:
                    seconds = DURATION_SECONDS.get(light.duration, light.duration * 300)
                    minutes = seconds // 60
                    duration_str = f"~{minutes} min"
                    
                    if "lantern" in name.lower():
                        notes = "Brightest portable light, reusable"
                    elif "torch" in name.lower():
                        notes = "Standard portable light"
                    elif "candle" in name.lower():
                        notes = "Dim light, burns quickly"
                    else:
                        notes = ""
                
                values = ["Item", name, light_level, duration_str, notes]
                self._add_row(ws, row, values, row % 2 == 0)
                row += 1
        
        # Add separator
        self._add_row(ws, row, ["", "", "", "", ""], False)
        row += 1
        
        # Light spells with proper circle info and light levels from constants
        # Spell durations scale with caster level; base duration ~2-3 minutes per circle
        # Only showing player-castable spells (Sunlight is NPC-only internal variant)
        light_spells = [
            ("Light", 1, "~3 min", "IN LOR - Basic illumination"),
            ("Night Vision", 3, "~6 min", "QUAS LOR - See in darkness"),
            ("Daylight", 6, "~10 min", "VAS IN LOR - Bright area light"),
        ]
        
        for spell_name, circle, duration, notes in light_spells:
            light_level = LIGHT_SPELL_LEVELS.get(spell_name, circle)
            type_str = f"Spell (Circle {circle})"
            values = [type_str, spell_name, light_level, duration, notes]
            self._add_row(ws, row, values, row % 2 == 0)
            row += 1
        
        self._auto_column_width(ws)
    
    def export_npc_names(self, npc_names: Dict, strings_parser=None) -> None:
        """Export NPC names (filtering out non-NPC strings from block 7)."""
        headers = ["Conv Slot", "NPC Name"]
        ws = self._create_sheet("NPC Names", headers)
        
        row = 2
        for slot in sorted(npc_names.keys()):
            name = npc_names[slot]
            # Skip entries that are not actual NPC names
            if not name or not name.strip():
                continue
            if "deal" in name.lower():
                continue
            if name.startswith("..."):
                continue
            if "cannot talk" in name.lower() or "no response" in name.lower():
                continue
            if "that I am getting" in name:
                continue
            
            self._add_row(ws, row, [slot, name.strip()], row % 2 == 0)
            row += 1
        self._auto_column_width(ws)
    
    def save(self, filename: str = "ultima_underworld_data.xlsx") -> Path:
        """Save the workbook."""
        filepath = self.output_path / filename
        self.wb.save(filepath)
        return filepath
